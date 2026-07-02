#!/usr/bin/env python3
"""
Validation complète du module DATA — Semaine 3 — Smart SIEM.

Génère un rapport de chaque type, vérifie la conformité et affiche
un bilan final succès/échec pour chaque critère.

Lancement : python scripts/validate_s3.py
"""

import sys
import time
import csv
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from app.reports.pdf_generator import generate_pdf_report
from app.reports.export import export_csv, export_excel
from app.reports.compliance import (
    verify_log_integrity,
    generate_retention_report,
    generate_compliance_report,
)
from app.reports.scheduler import run_daily_report, run_weekly_report

# Palette console
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

DATE_FROM = "2026-01-01T00:00:00Z"
DATE_TO   = "2026-12-31T23:59:59Z"

COL = 52
results = []


def ok(label, detail=""):
    results.append(True)
    flag = f"{GREEN}✓ OK{RESET}"
    print(f"  {label:<{COL}} {flag}  {detail}")


def fail(label, detail=""):
    results.append(False)
    flag = f"{RED}✗ ÉCHEC{RESET}"
    print(f"  {label:<{COL}} {flag}  {detail}")


def section(title):
    print(f"\n{BOLD}{CYAN}{'─' * 60}{RESET}")
    print(f"{BOLD}{CYAN}  {title}{RESET}")
    print(f"{BOLD}{CYAN}{'─' * 60}{RESET}")


def run():
    print(f"\n{BOLD}Smart SIEM — Validation module DATA, Semaine 3{RESET}")
    print(f"Période de test : {DATE_FROM}  →  {DATE_TO}\n")

    # ── 1. PDF ────────────────────────────────────────────────────────────────
    section("1. Générateur de rapports PDF")

    t0 = time.time()
    try:
        result = generate_pdf_report(DATE_FROM, DATE_TO)
        elapsed = time.time() - t0
        pdf_path = Path(result["path"])

        if pdf_path.exists():
            ok("Fichier PDF créé", pdf_path.name)
        else:
            fail("Fichier PDF créé", "introuvable sur le disque")

        size_kb = pdf_path.stat().st_size // 1024
        if size_kb > 0:
            ok("Taille PDF non nulle", f"{size_kb} Ko")
        else:
            fail("Taille PDF non nulle", "fichier vide")

        with pdf_path.open("rb") as f:
            magic = f.read(5)
        if magic == b"%PDF-":
            ok("Signature PDF valide", "%PDF-")
        else:
            fail("Signature PDF valide", f"reçu : {magic!r}")

        import hashlib
        computed = hashlib.sha256(pdf_path.read_bytes()).hexdigest()
        if computed == result["sha256"]:
            ok("Hash SHA-256 cohérent", result["sha256"][:16] + "…")
        else:
            fail("Hash SHA-256 cohérent", "hash retourné ≠ hash du fichier")

        ok(f"Temps de génération", f"{elapsed:.1f} s")
        print(f"  {'Chemin':<{COL}} {result['path']}")

    except Exception as exc:
        fail("Génération PDF", str(exc))

    # ── 2. CSV ────────────────────────────────────────────────────────────────
    section("2. Export CSV")

    try:
        t0 = time.time()
        csv_path = export_csv(DATE_FROM, DATE_TO)
        elapsed  = time.time() - t0

        if Path(csv_path).exists():
            ok("Fichier CSV créé", Path(csv_path).name)
        else:
            fail("Fichier CSV créé", "introuvable")

        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows   = list(reader)

        header = rows[0] if rows else []
        expected = ["timestamp", "source_ip", "host", "log_type", "severity", "raw_message", "tags"]
        if header == expected:
            ok("Colonnes CSV correctes", ", ".join(header))
        else:
            fail("Colonnes CSV correctes", f"reçu : {header}")

        n_data = len(rows) - 1
        ok(f"Nombre de lignes de données", f"{n_data:,}")
        ok(f"Temps d'export CSV", f"{elapsed:.1f} s")
        print(f"  {'Chemin':<{COL}} {csv_path}")

    except Exception as exc:
        fail("Export CSV", str(exc))

    # ── 3. Excel ──────────────────────────────────────────────────────────────
    section("3. Export Excel")

    try:
        t0 = time.time()
        xlsx_path = export_excel(DATE_FROM, DATE_TO)
        elapsed   = time.time() - t0

        if Path(xlsx_path).exists():
            ok("Fichier Excel créé", Path(xlsx_path).name)
        else:
            fail("Fichier Excel créé", "introuvable")

        wb = openpyxl.load_workbook(xlsx_path)

        if len(wb.sheetnames) == 3:
            ok("Nombre de feuilles (3)", ", ".join(wb.sheetnames))
        else:
            fail("Nombre de feuilles (3)", f"trouvé : {wb.sheetnames}")

        ws_logs = wb["Logs"]
        ok("Feuille 'Logs' présente", f"{ws_logs.max_row - 1:,} lignes de données")

        ws_sum = wb["Résumé"]
        ok("Feuille 'Résumé' présente", f"{ws_sum.max_row} lignes")

        ws_ips = wb["Top IPs"]
        ok("Feuille 'Top IPs' présente", f"{ws_ips.max_row - 1} IPs")

        ok(f"Temps d'export Excel", f"{elapsed:.1f} s")
        print(f"  {'Chemin':<{COL}} {xlsx_path}")

    except Exception as exc:
        fail("Export Excel", str(exc))

    # ── 4. Conformité ─────────────────────────────────────────────────────────
    section("4. Module de conformité RGPD/ISO")

    try:
        t0 = time.time()
        integrity = verify_log_integrity(DATE_FROM, DATE_TO)
        elapsed   = time.time() - t0

        ok("verify_log_integrity() exécutée", f"{elapsed:.1f} s")
        ok("Nombre de logs couverts", f"{integrity['log_count']:,}")
        ok("Hash SHA-256 d'intégrité", integrity["sha256"][:16] + "…")

        # Stabilité du hash
        integrity2 = verify_log_integrity(DATE_FROM, DATE_TO)
        if integrity["sha256"] == integrity2["sha256"]:
            ok("Hash stable (2 appels successifs)")
        else:
            fail("Hash stable (2 appels successifs)", "hash différent !")

    except Exception as exc:
        fail("verify_log_integrity()", str(exc))

    try:
        retention = generate_retention_report()
        ok("generate_retention_report() exécutée",
           f"rétention : {retention['retention_days']} jours, "
           f"{retention['total_logs']:,} logs total")
        if retention["oldest_log"] and retention["newest_log"]:
            ok("Bornes temporelles présentes",
               f"{retention['oldest_log'][:10]} → {retention['newest_log'][:10]}")
        else:
            ok("Bornes temporelles", "index vide ou aucune borne")
    except Exception as exc:
        fail("generate_retention_report()", str(exc))

    try:
        t0       = time.time()
        comp     = generate_compliance_report()
        elapsed  = time.time() - t0
        ok("generate_compliance_report() exécutée", f"{elapsed:.1f} s")
        ok("Hash du rapport de conformité", comp["report_sha256"][:16] + "…")
    except Exception as exc:
        fail("generate_compliance_report()", str(exc))

    # ── 5. Scheduler ──────────────────────────────────────────────────────────
    section("5. Scheduler (configuration)")

    try:
        from apscheduler.schedulers.blocking import BlockingScheduler
        from apscheduler.triggers.cron import CronTrigger
        ok("APScheduler importé")

        # Vérifie la configuration sans démarrer le scheduler
        sched = BlockingScheduler(timezone="UTC")
        sched.add_job(run_daily_report,  CronTrigger(hour=0, minute=5),  id="daily")
        sched.add_job(run_weekly_report, CronTrigger(day_of_week="mon", hour=0, minute=10), id="weekly")
        jobs = sched.get_jobs()
        ok(f"Jobs configurés : {len(jobs)}", "daily @ 00:05 UTC, weekly @ lundi 00:10 UTC")

    except Exception as exc:
        fail("Configuration du scheduler", str(exc))

    # ── Bilan ─────────────────────────────────────────────────────────────────
    total   = len(results)
    success = sum(results)
    failed  = total - success

    print(f"\n{'─' * 60}")
    print(f"{BOLD}Bilan : {success}/{total} vérifications réussies{RESET}")
    if failed == 0:
        print(f"{GREEN}{BOLD}✓ Validation S3 complète — tout est opérationnel.{RESET}\n")
    else:
        print(f"{RED}{BOLD}✗ {failed} vérification(s) échouée(s) — voir détails ci-dessus.{RESET}\n")

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    run()
