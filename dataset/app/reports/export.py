"""
Export CSV et Excel — Smart SIEM Module DATA, Semaine 3.

Importable par le Backend :
    from app.reports.export import export_csv, export_excel
"""

import csv
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

_BASE_DIR = Path(__file__).resolve().parent.parent.parent
load_dotenv(_BASE_DIR / ".env")
sys.path.insert(0, str(_BASE_DIR))

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.aggregations import count_by_log_type, count_by_severity, top_source_ips
from app.db.search import search_logs

_EXPORTS_DIR = Path(os.getenv("EXPORTS_DIR", str(_BASE_DIR / "exports")))

_LOG_COLUMNS = ["timestamp", "source_ip", "host", "log_type", "severity", "raw_message", "tags"]

# Taille de page utilisée pour paginer les appels à search_logs
_PAGE_SIZE = 1000


def _fetch_all_logs(date_from=None, date_to=None, severity=None, log_type=None):
    """
    Récupère tous les logs correspondant aux filtres via search_logs() paginé.

    search_logs() n'a pas de limite supérieure sur page_size — on pagine
    par blocs de _PAGE_SIZE pour gérer les grands volumes sans saturer ES.
    """
    page = 1
    all_results = []
    while True:
        resp = search_logs(
            date_from=date_from,
            date_to=date_to,
            severity=severity,
            log_type=log_type,
            page=page,
            page_size=_PAGE_SIZE,
        )
        batch = resp["results"]
        all_results.extend(batch)
        if len(all_results) >= resp["total"] or not batch:
            break
        page += 1
    return all_results


def _default_export_path(suffix):
    """Retourne un chemin exports/logs_{timestamp}.{suffix}."""
    _EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return _EXPORTS_DIR / f"logs_{ts}.{suffix}"


def _row_from_log(log):
    """Extrait les colonnes exportables d'un document ES."""
    tags = log.get("tags", [])
    if isinstance(tags, list):
        tags = ", ".join(tags)
    return [
        log.get("timestamp", ""),
        log.get("source_ip", ""),
        log.get("host", ""),
        log.get("log_type", ""),
        log.get("severity", ""),
        log.get("raw_message", ""),
        tags,
    ]


# ── CSV ──────────────────────────────────────────────────────────────────────

def export_csv(
    date_from=None,
    date_to=None,
    severity=None,
    log_type=None,
    output_path=None,
):
    """
    Exporte les logs filtrés en CSV.

    Colonnes : timestamp, source_ip, host, log_type, severity,
               raw_message, tags.

    Les logs sont récupérés via search_logs() paginé et écrits en une
    passe dans le fichier CSV (encodage UTF-8 avec BOM pour compatibilité Excel).

    Paramètres
    ----------
    date_from   : str, optionnel — borne inférieure ISO 8601.
    date_to     : str, optionnel — borne supérieure ISO 8601.
    severity    : str, optionnel — filtre sévérité (info/warning/critical).
    log_type    : str, optionnel — filtre type (auth/réseau/système/application).
    output_path : str | Path, optionnel — chemin de sortie.
                  Par défaut : exports/logs_{timestamp}.csv

    Retour
    ------
    str — chemin absolu du fichier CSV généré.

    Exemple
    -------
    >>> path = export_csv(severity="critical", date_from="2026-06-01T00:00:00Z")
    >>> print(path)
    /…/exports/logs_20260623T120000Z.csv
    """
    if output_path is None:
        output_path = _default_export_path("csv")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logs = _fetch_all_logs(
        date_from=date_from,
        date_to=date_to,
        severity=severity,
        log_type=log_type,
    )

    # utf-8-sig : ajoute le BOM pour que Excel ouvre correctement l'UTF-8
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(_LOG_COLUMNS)
        for log in logs:
            writer.writerow(_row_from_log(log))

    return str(output_path.resolve())


# ── Excel ─────────────────────────────────────────────────────────────────────

def _style_header_row(ws, row=1, fill_hex="0D2035", font_hex="1F6FEB"):
    """Applique un style d'en-tête foncé à la première ligne d'une feuille."""
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_hex)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(ws, min_width=10, max_width=60):
    """Ajuste automatiquement la largeur des colonnes au contenu."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def export_excel(
    date_from=None,
    date_to=None,
    severity=None,
    log_type=None,
    output_path=None,
):
    """
    Exporte les logs filtrés en Excel (.xlsx) avec trois feuilles.

    Feuille 1 "Logs"    : tous les logs filtrés (colonnes identiques au CSV).
    Feuille 2 "Résumé"  : count_by_severity() et count_by_log_type().
    Feuille 3 "Top IPs" : top_source_ips(n=20).

    Paramètres
    ----------
    date_from   : str, optionnel — borne inférieure ISO 8601.
    date_to     : str, optionnel — borne supérieure ISO 8601.
    severity    : str, optionnel — filtre sévérité.
    log_type    : str, optionnel — filtre type.
    output_path : str | Path, optionnel — chemin de sortie.
                  Par défaut : exports/logs_{timestamp}.xlsx

    Retour
    ------
    str — chemin absolu du fichier Excel généré.

    Exemple
    -------
    >>> path = export_excel(date_from="2026-06-01T00:00:00Z", date_to="2026-06-07T23:59:59Z")
    >>> print(path)
    /…/exports/logs_20260623T120000Z.xlsx
    """
    if output_path is None:
        output_path = _default_export_path("xlsx")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logs           = _fetch_all_logs(date_from=date_from, date_to=date_to, severity=severity, log_type=log_type)
    sev_counts     = count_by_severity(date_from=date_from, date_to=date_to)
    type_counts    = count_by_log_type(date_from=date_from, date_to=date_to)
    top_ips        = top_source_ips(n=20, date_from=date_from, date_to=date_to)

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Logs ──────────────────────────────────────────────────
    ws_logs = wb.active
    ws_logs.title = "Logs"
    ws_logs.append(_LOG_COLUMNS)
    for log in logs:
        ws_logs.append(_row_from_log(log))
    _style_header_row(ws_logs)
    _autofit_columns(ws_logs)

    # ── Feuille 2 : Résumé ────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Résumé")
    ws_sum.append(["Sévérité", "Nombre"])
    for key in ["critical", "warning", "info"]:
        ws_sum.append([key, sev_counts.get(key, 0)])
    ws_sum.append([])
    ws_sum.append(["Type de log", "Nombre"])
    for key in ["auth", "réseau", "système", "application"]:
        ws_sum.append([key, type_counts.get(key, 0)])
    _style_header_row(ws_sum, row=1)
    _style_header_row(ws_sum, row=5)
    _autofit_columns(ws_sum)

    # ── Feuille 3 : Top IPs ───────────────────────────────────────────────
    ws_ips = wb.create_sheet("Top IPs")
    ws_ips.append(["#", "Adresse IP", "Nombre de logs"])
    for i, entry in enumerate(top_ips, 1):
        ws_ips.append([i, entry["ip"], entry["count"]])
    _style_header_row(ws_ips)
    _autofit_columns(ws_ips)

    wb.save(str(output_path))

    return str(output_path.resolve())
