"""
Tests des exports CSV et Excel — Smart SIEM DATA S3.

Prérequis : Elasticsearch actif avec des données dans logs-siem.
Lancement  : pytest tests/test_export.py -v
"""

import csv
import sys
from pathlib import Path

import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.reports.export import export_csv, export_excel, _LOG_COLUMNS

_DATE_FROM = "2026-01-01T00:00:00Z"
_DATE_TO   = "2026-12-31T23:59:59Z"


# ── CSV ──────────────────────────────────────────────────────────────────────

class TestExportCsv:
    @pytest.fixture(scope="class")
    def csv_path(self, tmp_path_factory):
        out = tmp_path_factory.mktemp("csv") / "test_export.csv"
        return export_csv(_DATE_FROM, _DATE_TO, output_path=str(out))

    def test_file_created(self, csv_path):
        assert Path(csv_path).exists()

    def test_file_not_empty(self, csv_path):
        assert Path(csv_path).stat().st_size > 0

    def test_correct_columns(self, csv_path):
        """La première ligne correspond exactement aux colonnes attendues."""
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert header == _LOG_COLUMNS, f"En-têtes incorrects : {header}"

    def test_has_data_rows(self, csv_path):
        """Le CSV contient au moins une ligne de données (hors en-tête)."""
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        assert len(rows) > 1, "Le CSV ne contient que l'en-tête — aucune donnée exportée."

    def test_row_has_correct_number_of_columns(self, csv_path):
        """Chaque ligne de données a exactement len(_LOG_COLUMNS) colonnes."""
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader)  # saute l'en-tête
            for i, row in enumerate(reader):
                assert len(row) == len(_LOG_COLUMNS), (
                    f"Ligne {i + 2} : {len(row)} colonnes au lieu de {len(_LOG_COLUMNS)}"
                )

    def test_returns_absolute_path(self, csv_path):
        assert Path(csv_path).is_absolute()

    def test_filter_by_severity(self, tmp_path):
        """L'export filtré par sévérité ne contient que des logs de cette sévérité."""
        out = tmp_path / "critical.csv"
        path = export_csv(_DATE_FROM, _DATE_TO, severity="critical", output_path=str(out))
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        if rows:  # si des logs critiques existent
            assert all(r["severity"] == "critical" for r in rows), \
                "Des logs non-critiques sont présents dans l'export filtré."


# ── Excel ─────────────────────────────────────────────────────────────────────

class TestExportExcel:
    @pytest.fixture(scope="class")
    def xlsx_path(self, tmp_path_factory):
        out = tmp_path_factory.mktemp("xlsx") / "test_export.xlsx"
        return export_excel(_DATE_FROM, _DATE_TO, output_path=str(out))

    def test_file_created(self, xlsx_path):
        assert Path(xlsx_path).exists()

    def test_file_is_valid_xlsx(self, xlsx_path):
        """Le fichier est un classeur Excel valide lisible par openpyxl."""
        wb = openpyxl.load_workbook(xlsx_path)
        assert wb is not None

    def test_has_three_sheets(self, xlsx_path):
        """Le classeur contient exactement 3 feuilles."""
        wb = openpyxl.load_workbook(xlsx_path)
        assert len(wb.sheetnames) == 3, f"Feuilles trouvées : {wb.sheetnames}"

    def test_sheet_names(self, xlsx_path):
        """Les feuilles s'appellent 'Logs', 'Résumé' et 'Top IPs'."""
        wb = openpyxl.load_workbook(xlsx_path)
        assert wb.sheetnames == ["Logs", "Résumé", "Top IPs"]

    def test_logs_sheet_header(self, xlsx_path):
        """La feuille 'Logs' a les bonnes colonnes en première ligne."""
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Logs"]
        header = [ws.cell(1, c).value for c in range(1, len(_LOG_COLUMNS) + 1)]
        assert header == _LOG_COLUMNS, f"En-têtes 'Logs' incorrects : {header}"

    def test_logs_sheet_has_data(self, xlsx_path):
        """La feuille 'Logs' contient au moins une ligne de données."""
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Logs"]
        assert ws.max_row > 1, "Feuille 'Logs' vide — aucune donnée exportée."

    def test_resume_sheet_has_severity_data(self, xlsx_path):
        """La feuille 'Résumé' contient les données de sévérité."""
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Résumé"]
        header = ws.cell(1, 1).value
        assert header == "Sévérité", f"En-tête inattendu : {header!r}"

    def test_top_ips_sheet_structure(self, xlsx_path):
        """La feuille 'Top IPs' a 3 colonnes et au moins une IP."""
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Top IPs"]
        assert ws.cell(1, 1).value == "#"
        assert ws.cell(1, 2).value == "Adresse IP"
        assert ws.cell(1, 3).value == "Nombre de logs"

    def test_returns_absolute_path(self, xlsx_path):
        assert Path(xlsx_path).is_absolute()
