# backend/app/services/export_service.py
#
# Génération d'exports CSV et Excel (.xlsx) réels à partir de listes de
# dictionnaires (logs Elasticsearch ou alertes SQL). Aucune donnée fictive :
# ce module se contente de sérialiser ce qui lui est transmis.

import csv
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def to_csv_bytes(rows: list[dict], columns: list[str]) -> bytes:
    """Sérialise une liste de dicts en CSV (colonnes dans l'ordre fourni)."""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({col: row.get(col, "") for col in columns})
    return buffer.getvalue().encode("utf-8-sig")  # BOM pour compatibilité Excel


def to_xlsx_bytes(rows: list[dict], columns: list[str], sheet_title: str = "Export") -> bytes:
    """Sérialise une liste de dicts en classeur Excel (.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # limite Excel sur le nom de feuille

    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    # Largeur de colonne raisonnable basée sur le contenu
    for idx, col in enumerate(columns, start=1):
        max_len = max([len(str(col))] + [len(str(r.get(col, ""))) for r in rows[:200]])
        ws.column_dimensions[get_column_letter(idx)].width = min(60, max(10, max_len + 2))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
